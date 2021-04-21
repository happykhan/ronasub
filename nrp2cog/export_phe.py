import gspread
import logging 
import paramiko
import os 
from openpyxl import Workbook

class PheFiles():

    def __init__(self, file_server, username, key):
        ssh = paramiko.SSHClient()
        the_key = paramiko.RSAKey.from_private_key_file(key)
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(file_server, username=username, pkey=the_key, port=443)
        self.sftp = ssh.open_sftp() 

    def create_dir(self, dir_name):
        try:
            self.sftp.stat(dir_name)
        except FileNotFoundError:
            logging.debug(f'dir {dir_name} on remote not found, creating')
            self.sftp.mkdir(dir_name)
        return dir_name

    def put_file(self, filename, path, overwrite=False):
        remote_file_path = os.path.join(path, os.path.basename(filename))
        if overwrite:
            logging.debug(f'sending file {filename}')
            self.sftp.remove(remote_file_path)
            self.sftp.put(filename, remote_file_path) 
        else:
            try:
                self.sftp.stat(remote_file_path)
            except FileNotFoundError:
                logging.debug(f'sending file {filename}')
                self.sftp.put(filename, remote_file_path) 
        return remote_file_path


def export_phe(client, out_dir, sheet_name, file_server, username, key):
    sheet = client.open(sheet_name).sheet1
    all_values = sheet.get_all_records()
    load = { }
    logging.info('Reading values from master table')
    for x in all_values:
        date = ''
        if len(x['collection_date']) > 3 :
            date = x['collection_date']
        else:
            date = x['received_date']
        if not x['central_sample_id'].endswith('duplicate') and not x['central_sample_id'].endswith('duplicates') and len(x["biosample_source_id"]) > 6:
            load[x['central_sample_id']] = { "cog_id": x['central_sample_id'], "biosample_source_id": x["biosample_source_id"], "date" : date}
    out_file = os.path.join(out_dir, f'NORW-{sheet_name}-cog2labid.xlsx') 
    """
    FORMAT:
    Local laboratory ID	COG-UK sample ID	Date of specimen
		
    e.g. LAB99999	e.g. 987654321	DD/MM/YYYY
    """

    if not os.path.exists(out_dir):
        os.mkdir(out_dir)
    logging.info('Writing to temp file')
    wb = Workbook()
    ws1 = wb.active
    ws1.title = 'NORW'
    ws1.cell(column=1, row=1, value='Local laboratory ID')
    ws1.cell(column=2, row=1, value='COG-UK sample ID')
    ws1.cell(column=3, row=1, value='Date of specimen')
    row_count = 2
    for record in load.values():

        ws1.cell(column=1, row=row_count,value=record['biosample_source_id'])
        ws1.cell(column=2, row=row_count,value=record['cog_id'])
        ws1.cell(column=3, row=row_count,value=record['date'])
        row_count += 1 

    wb.save(filename = out_file)

    logging.info('Uploading to server')        
    server = PheFiles(file_server, username, key)
    server.put_file(out_file, 'upload', overwrite=False)
    logging.info('Done')                    
